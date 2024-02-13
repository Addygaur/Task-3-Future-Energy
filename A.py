import io
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader

def excel_to_pdf(input_excel, output_pdf):
    # Load the Excel workbook
    wb = load_workbook(filename=input_excel)

    # Create a PDF canvas
    c = canvas.Canvas(output_pdf, pagesize=letter)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Extract data from the sheet
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)

        # Define table style
        style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)])

        # Create table
        table = Table(data)
        table.setStyle(style)

        # Get table width and height
        width, height = letter
        table.wrapOn(c, width, height)
        table_width, table_height = table.wrap(width, height)

        # Draw table on the canvas
        table.drawOn(c, 50, height - table_height - 50)

        # Check if there's an image below the table
        for row_idx, row in enumerate(ws.iter_rows()):
            for col_idx, cell in enumerate(row):
                if col_idx == len(row) - 1:  # Check if it's the last column
                    if cell._value and hasattr(cell._value, "anchor"):
                        anchor = cell._value.anchor
                        if hasattr(anchor, '_image'):
                            image = anchor._image
                            image_data = image._data
                            img = ImageReader(io.BytesIO(image_data))
                            c.drawImage(img, 50, height - table_height - 150, width=200, height=200)
                            break  # Only print the first image below the table
            else:
                continue
            break

        # Add a page break for the next sheet
        c.showPage()

    # Save the PDF
    c.save()

# Input and output file paths
input_excel_path = "InputXls.xlsx"
output_pdf_path = "Output-PDF.pdf"

# Convert Excel to PDF
excel_to_pdf(input_excel_path, output_pdf_path)
